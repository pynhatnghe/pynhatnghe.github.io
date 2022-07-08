from openpyxl import Workbook
import random
wb = Workbook()

ws = wb.active # Lấy cái workshet đang active (sheet đầu tiên)

# Tạo dữ liệu để vẽ chart
ws.cell(row=1, column=3, value="BẢNG DỮ LIỆU")
# Tạo dữ liệu random từ dòng 2 đến dòng 5
for dong in range(2,6):
    for cot in range(1, 6):
        ws.cell(
            row=dong, column=cot,
            value=random.randint(1,55)
        )

# Vẽ chart
from openpyxl.chart import BarChart, Reference, Series
values = Reference(
    ws, min_col=1, max_col=5, min_row=2, max_row=5
)
bar_chart = BarChart()
bar_chart.add_data(values)

# Chèn chart vào vị trí nào
ws.add_chart(bar_chart, "A7")

wb.save("ChartWithDataRandom.xlsx")