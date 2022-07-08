from openpyxl import load_workbook

wb = load_workbook("ChartWithDataRandom.xlsx")
print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]

# Đọc những ô có giá trị
for row in ws.values:
    for value in row:
        print(value, "\t", end='')
    print("")