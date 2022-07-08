from openpyxl import Workbook
# T5ao workwork - file excel
wb = Workbook()
# Tạo worksheet
ws = wb.create_sheet("ThongTin")

# Gán giá trị từng ô (cell)
ws["A1"] = "Họ tên"
ws["B1"] = "Số điện thoại"

# Gán giá trị cho từng dòng
ws.append(["Trần Văn Tèo", "0909009900"])

for dong in range(3, 21):
    for cot in range(1, 21):
        ws.cell(row=dong, column=cot, value=f"{dong}x{cot}")

# Save
wb.save("DemoExcel.xlsx")

