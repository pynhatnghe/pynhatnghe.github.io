# pip install openpyxl
# File excel (Workbook), trong file sẽ có nhiều Worksheet, trong worksheet có nhiều cell
from openpyxl import Workbook

wb = Workbook()

# Tạo worksheet có tên NhatNghe
ws = wb.create_sheet("NhatNghe", 1)

# Ghi ô "A1"
ws["A1"] = "Trung Tâm Nhất Nghệ"

# Thêm dòng mới
ws.append([2, 3, 4])

# Lưu file
wb.save("DemoOpenpyxl.xlsx")
